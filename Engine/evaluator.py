import os
import openpyxl
from Methods import gpt_request, stimy_request


class MathEvaluation:
    def __init__(self, path_to_create, path_to_data):
        self.path_to_create = path_to_create
        self.path_to_data = path_to_data
        self.sheet_names = ["problems", "gpt_answers", "ev-gpt->stimy", "stimy_answers", "ev-stimy->gpt"]
        self.create_excel_file()

    def create_excel_file(self):
        if not os.path.exists(self.path_to_create):
            os.makedirs(self.path_to_create)

        workbook = openpyxl.Workbook()
        for i, sheet_name in enumerate(self.sheet_names):
            if i == 0:
                workbook.active.title = sheet_name
            else:
                workbook.create_sheet(title=sheet_name)
        workbook.save(self.path_to_data)

    def add_problems(self):
        workbook = openpyxl.load_workbook(self.path_to_data)
        sheet = workbook["problems"]

        stimy_problems = [stimy_request.get_answer("Generate a high school math problem") for _ in range(2)]
        gpt_problems = [gpt_request.gpt_message("Generate a high school math problem") for _ in range(2)]

        for idx, problem in enumerate(stimy_problems + gpt_problems, start=1):
            sheet.cell(row=idx, column=1, value=problem)

        workbook.save(self.path_to_data)

    def generate_answers(self):
        workbook = openpyxl.load_workbook(self.path_to_data)
        problems_sheet = workbook["problems"]
        gpt_answers_sheet = workbook["gpt_answers"]
        stimy_answers_sheet = workbook["stimy_answers"]

        for idx, row in enumerate(problems_sheet.iter_rows(values_only=True), start=1):
            problem = row[0]

            gpt_answer = gpt_request.gpt_message(problem)
            stimy_answer = stimy_request.get_answer(problem)

            gpt_answers_sheet.cell(row=idx, column=1, value=gpt_answer)
            stimy_answers_sheet.cell(row=idx, column=1, value=stimy_answer)

        workbook.save(self.path_to_data)

    def evaluate_answers(self):
        workbook = openpyxl.load_workbook(self.path_to_data)
        problems_sheet = workbook["problems"]
        gpt_answers_sheet = workbook["gpt_answers"]
        stimy_answers_sheet = workbook["stimy_answers"]
        ev_gpt_stimy_sheet = workbook["ev-gpt->stimy"]
        ev_stimy_gpt_sheet = workbook["ev-stimy->gpt"]

        for idx, row in enumerate(problems_sheet.iter_rows(values_only=True), start=1):
            problem = row[0]
            gpt_answer = gpt_answers_sheet.cell(row=idx, column=1).value
            stimy_answer = stimy_answers_sheet.cell(row=idx, column=1).value

            gpt_eval = gpt_request.gpt_message(
                f"Rate the solution quality for the problem: {problem} and solution: {stimy_answer} on a scale of 1-5.")
            stimy_eval = stimy_request.get_answer(
                f"Rate the solution quality for the problem: {problem} and solution: {gpt_answer} on a scale of 1-5.")

            ev_gpt_stimy_sheet.cell(row=idx, column=1, value=gpt_eval)
            ev_stimy_gpt_sheet.cell(row=idx, column=1, value=stimy_eval)

            print(f"Problem {idx} -> Stimy Eval: {stimy_eval}, GPT Eval: {gpt_eval}")

        workbook.save(self.path_to_data)



