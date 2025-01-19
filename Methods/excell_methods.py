from openpyxl import load_workbook, Workbook


class ExcelManager:
    def __init__(self, file_path, sheet_name=None):
        """
        Inicjalizuje obiekt ExcelManager i ładuje istniejący plik Excel lub tworzy nowy.

        :param file_path: Ścieżka do pliku Excel
        :param sheet_name: Nazwa arkusza, na którym będą wykonywane operacje
        """
        self.file_path = file_path
        try:
            self.workbook = load_workbook(file_path)
            if sheet_name:
                if sheet_name in self.workbook.sheetnames:
                    self.sheet = self.workbook[sheet_name]
                else:
                    raise ValueError(f"Arkusz '{sheet_name}' nie istnieje w pliku '{file_path}'.")
            else:
                self.sheet = self.workbook.active
        except FileNotFoundError:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            if sheet_name:
                self.sheet.title = sheet_name
            self.save_workbook()

    def save_workbook(self):
        """Zapisuje zmiany w pliku Excel."""
        self.workbook.save(self.file_path)

    def read_cell(self, cell_address):
        """
        Odczytuje wartość z konkretnej komórki.

        :param cell_address: Adres komórki, np. "A1"
        :return: Wartość w komórce
        """
        return self.sheet[cell_address].value

    def write_cell(self, cell_address, value):
        """
        Zapisuje wartość do konkretnej komórki.

        :param cell_address: Adres komórki, np. "A1"
        :param value: Wartość do zapisania
        """
        self.sheet[cell_address] = value
        self.save_workbook()

    def read_row(self, row_number):
        """
        Odczytuje wszystkie wartości z wiersza.

        :param row_number: Numer wiersza
        :return: Lista wartości w wierszu
        """
        return [cell.value for cell in self.sheet[row_number]]

    def read_column(self, column_letter):
        """
        Odczytuje wszystkie wartości z kolumny.

        :param column_letter: Litera kolumny, np. "A"
        :return: Lista wartości w kolumnie
        """
        return [cell.value for cell in self.sheet[column_letter]]

    def add_sheet(self, sheet_name):
        """
        Dodaje nowy arkusz do pliku Excel.

        :param sheet_name: Nazwa nowego arkusza
        """
        if sheet_name in self.workbook.sheetnames:
            raise ValueError(f"Arkusz '{sheet_name}' już istnieje.")
        self.workbook.create_sheet(title=sheet_name)
        self.save_workbook()

    def delete_sheet(self, sheet_name):
        """
        Usuwa istniejący arkusz z pliku Excel.

        :param sheet_name: Nazwa arkusza do usunięcia
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Arkusz '{sheet_name}' nie istnieje.")
        sheet_to_delete = self.workbook[sheet_name]
        self.workbook.remove(sheet_to_delete)
        self.save_workbook()

