import os
from openpyxl import Workbook

def create_excel_file(folder_path, file_name, sheet_names=None):
    """
    :param folder_path: Ścieżka do folderu, gdzie plik ma zostać zapisany
    :param file_name: Nazwa pliku Excel (bez rozszerzenia)
    :param sheet_names: Lista nazw arkuszy do dodania do pliku Excel
    """
    os.makedirs(folder_path, exist_ok=True)

    file_path = os.path.join(folder_path, f"{file_name}.xlsx")

    if not os.path.isfile(file_path):
        workbook = Workbook()

        if sheet_names:
            if "Sheet" in workbook.sheetnames:
                del workbook["Sheet"]
            for name in sheet_names:
                workbook.create_sheet(title=name)

        workbook.save(file_path)
        print(f"Plik Excel '{file_name}.xlsx' został utworzony w folderze '{folder_path}'.")
    else:
        print(f"Plik '{file_name}.xlsx' już istnieje w folderze '{folder_path}'.")




